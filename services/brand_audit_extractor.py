"""
Brand Audit Extractor (Universal)
3-layer extraction: Guaranteed fields -> Auto-discover metafields -> Fallback parsing.
Works consistently across ANY Shopify store.

Usage:
    # Shopify only (public store)
    python brand_audit_extractor.py https://www.skinq.com/products.json
    python brand_audit_extractor.py https://www.skinq.com/products.json --output ./sheets/

    # Shopify Admin API with access token
    python brand_audit_extractor.py https://skinq.myshopify.com/admin \
        --access-token shpat_xxxxxxxxxxxxxxxxxxxx \
        --output ./sheets/

    # Shopify Admin API with Client ID + Secret
    python brand_audit_extractor.py https://skinq.myshopify.com/admin \
        --client-id YOUR_CLIENT_ID \
        --client-secret YOUR_CLIENT_SECRET \
        --output ./sheets/

    # Shopify + GA4 + GSC
    python brand_audit_extractor.py https://store.myshopify.com/admin \
        --access-token shpat_xxxxxxxxxxxxxxxxxxxx \
        --ga4-property 526018830 \
        --ga4-credentials ./ga4_credentials.json \
        --gsc-site "https://www.example.com/" \
        --gsc-credentials ./ga4_credentials.json \
        --output ./sheets/
"""

import requests
import openpyxl
import re
import sys
import os
import json
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")


# ─── Audit Field Configuration ───────────────────────────────────────────────

AUDIT_FIELDS = [
    # ── Basic Product Info ───────────────────────────────────────────────────────
    ("Product Name",           "Display name in report",                                              "Validate"),
    ("PDP URL (Clean)",        "pdp_url -> Machine Readiness audit",                                  "Confirm canonical URL"),
    ("Product Subtitle",       "Short tagline / product sub title",                                   "Validate"),
    ("Selling Price (INR)",    "Class B: {price_range} -- CRITICAL for pricing hallucination fix",   "Validate -- actual selling price?"),
    ("MRP / Compare-at (INR)", "Discount context",                                                   "Confirm"),
    ("Discount %",             "Price positioning signal",                                             "--"),
    ("SKU / EAN",              "Product identifiers",                                                  "Validate"),
    ("Product Type",           "Class A: {product_type}",                                             "Validate"),
    ("Pack Size",              "Class B: {unit}",                                                     "FILL -- exact ml/g"),
    ("Total Inventory",        "Stock level",                                                         "--"),
    ("Category (Taxonomy)",    "Shopify standard taxonomy path",                                      "Confirm -- is category correct?"),
    # ── Ratings & Reviews ─────────────────────────────────────────────────────────
    ("Rating",                 "Brand Perception evidence (Judge.me)",                                "--"),
    ("Review Count",           "Social proof signal",                                                  "--"),
    ("Amazon Reviews",         "Cross-platform social proof",                                         "Confirm listing URL"),
    # ── Ingredients & Science ─────────────────────────────────────────────────────
    ("Hero Molecules",         "Class B: {attributes} -- feeds query templates",                    "Validate -- any key ingredients missing?"),
    ("Full Ingredient List",   "Complete INCI list from metafields",                                "Verify completeness"),
    ("Clinical Claims",        "Claim Compliance Framework -- must map to clinical trial source",    "CRITICAL -- provide study references"),
    ("Clinical Method",        "Evidence grounding",                                                   "FILL -- study name, sample size, duration"),
    ("How It Works",           "Mechanism / science behind the product",                             "Validate"),
    # ── Targeting ────────────────────────────────────────────────────────────────
    ("Target Skin Types",      "Class C: {specific_conditions} -- feeds condition queries",          "Validate -- any conditions missing?"),
    ("Who It's For",           "Detailed targeting from metafields",                                   "Validate"),
    ("Who It's NOT For",       "Contraindications",                                                   "FILL -- any safety concerns?"),
    # ── Trust & Authority ─────────────────────────────────────────────────────────
    ("Certifications",         "Trust signals for E-E-A-T",                                          "Confirm -- are certifications correct?"),
    ("Dr. Involvement",        "Authority signal for E-E-A-T",                                       "FILL -- did a formulator/expert create this product?"),
    ("Doctor Profiles",        "Doctor metaobject references",                                        "FILL -- doctor names + credentials"),
    # ── Product Experience ───────────────────────────────────────────────────────
    ("Results (Why You'll Love It)", "Claims and benefits listed on PDP",                            "Validate"),
    ("Texture / Experience",   "How the product feels on skin",                                      "Validate"),
    ("How To Use",             "Step-by-step usage instructions",                                     "Validate"),
    ("Shelf Life & Storage",   "Longevity and storage instructions",                                 "Confirm"),
    ("Before/After Images",    "Visuals from clinical photography",                                   "Confirm -- canKasparro use?"),
    # ── Shopability ──────────────────────────────────────────────────────────────
    ("Collections",            "Store categorization / collection membership",                        "--"),
    ("Related Products",        "Cross-sell opportunities",                                           "FILL -- validate recommendations"),
    # ── SEO & Site ───────────────────────────────────────────────────────────────
    ("SEO Title",              "Meta title",                                                          "Validate"),
    ("SEO Description",        "Meta description",                                                    "Validate"),
    ("Breadcrumb",             "Navigation breadcrumb path",                                          "Confirm"),
    ("Google Shopping Labels",  "mm-google-shopping custom labels",                                   "FILL -- product categorization"),
    # ── Marketing ────────────────────────────────────────────────────────────────
    ("Short Description",      "Product elevator pitch",                                               "Validate"),
    ("Product USPs",           "Unique selling propositions",                                         "Validate"),
    ("FAQ Content",            "Product-specific FAQs from PDP",                                      "Review"),
    # ── Admin ────────────────────────────────────────────────────────────────────
    ("Published Date",          "When product went live",                                              "--"),
    ("Price Positioning",      "CRITICAL -- how does brand position pricing?",                       "FILL -- budget, value-clinical, mid-range, or premium?"),
]


# ─── Layer 3: Fallback Constants ─────────────────────────────────────────────

MOLECULES = [
    "niacinamide", "salicylic acid", "hyaluronic acid", "vitamin c",
    "retinol", "azelaic acid", "glycerin", "ceramide", "squalane",
    "kojic acid", "alpha arbutin", "tea tree", "aloe vera",
    "collagen", "peptide", "zinc", "centella", "cica",
    "sodium hyaluronate", "evening primrose", "spf", "lactic acid",
    "glycolic acid", "benzoyl peroxide", "tranexamic acid",
    "argan oil", "rosehip oil", "bakuchiol", "malic acid",
]

SKIN_TYPE_KEYWORDS = {
    "dry": "Dry Skin", "oily": "Oily Skin", "sensitive": "Sensitive Skin",
    "acne": "Acne-Prone Skin", "normal": "Normal Skin",
    "dehydrated": "Dehydrated Skin", "all skin": "All Skin Types",
    "combination": "Combination Skin", "mature": "Mature Skin",
}

CERT_KEYWORDS = [
    "clinically proven", "dermatologist tested", "dermatologist recommended",
    "dermatologist formulated", "non-irritant", "cruelty free", "vegan",
    "organic", "natural", "fda approved", "iso certified", "made safe",
    "ecocert", "gmp certified", "paraben free", "sulfate free",
    "fragrance free", "derma",
]

# Keywords for mapping metafield definitions to audit fields
METAFIELD_MAPPING_RULES = {
    "hero_molecules": ["ingredient", "active_ingredient", "key_ingredient", "hero_molecule", "product_ingredients", "product_ingredient"],
    "full_ingredient_list": ["ingredient_list", "inci", "full_ingredient", "active_ingredients", "active_ingredient"],
    "target_skin_types": ["skin_type", "who_is_it_for", "suitable_for", "target_skin", "who_should_use"],
    "clinical_claims": ["clinical", "proven_result", "efficacy", "clinically_proven"],
    "how_it_works": ["how_it_works", "mechanism", "science_behind"],
    "certifications": ["certification", "cruelty", "vegan", "organic_cert"],
    "doctor_involvement": ["doctor", "formulator", "expert", "dermatologist", "physician", "doctor_profile"],
    "subtitle": ["subtitle", "sub_title", "short_description", "tagline", "product_sub"],
    "who_its_for": ["who_is_it_for", "suitable_for", "ideal_for", "best_for", "who_should"],
    "how_to_use": ["how_to_use", "direction", "usage", "application"],
    "pack_size": ["size", "volume", "weight", "net_weight", "quantity", "pack_size"],
    # Additional SkinQ-specific mappings
    "shelf_life": ["shelf_life", "dermatologist_recommended", "storage", "product_longevity"],
    "texture_experience": ["texture_experience"],
    "who_shouldnt": ["shouldn_t", "who_shouldn", "should_not", "not_suitable"],
    "results_why_love": ["results", "why_love", "why_choose", "why_choose_skinq"],
    "detailed_description": ["detailed_product_description"],
    "short_desc_field": ["short_description", "product_short_description"],
    "usps": ["usps", "product_usps"],
    "faq_content": ["faq", "product_faq", "faqs"],
    "ean_code": ["ean_code"],
    "breadcrumb": ["breadcrumb", "custom_breadcrumb"],
    "google_labels": ["custom_label"],
    "ingredient_bg_color": ["ingredient_background_color"],
}

# Review app detection order
REVIEW_APP_PATTERNS = [
    # (namespace, rating_key, count_key, rating_parser)
    ("reviews", "rating", "rating_count", "shopify_standard"),
    ("judgeme", "badge", None, "judgeme_html"),
    ("yotpo", "reviews_average", "reviews_count", "direct"),
    ("loox", "avg_rating", "num_reviews", "direct"),
    ("stamped", "reviews_average", "reviews_count", "direct"),
    ("automizely_reviews", "ratings", "raters", "direct"),
    ("reputon", None, None, "amazon_reviews"),
]

# Description heading classification
HEADING_CLASSIFICATION = {
    "INGREDIENTS": ["ingredient", "what's in it", "active", "formulation", "key ingredient"],
    "USAGE": ["how to use", "direction", "usage", "application", "how to apply"],
    "BENEFITS": ["benefit", "result", "why you'll love", "what it does"],
    "SKIN_TYPE": ["skin type", "suitable for", "best for", "who is it for", "ideal for"],
    "CLINICAL": ["clinical", "proven", "study", "tested", "research"],
    "CERTIFICATIONS": ["certification", "award", "accreditation", "safety"],
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def clean_html(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    text = text.replace("&quot;", '"').replace("&#39;", "'").replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", text).strip()


def get_discount(price: str, mrp: str) -> str:
    if not mrp or not price:
        return "—"
    try:
        p, m = float(price), float(mrp)
        if m > p:
            return f"{((m - p) / m * 100):.0f}%"
    except (ValueError, ZeroDivisionError):
        pass
    return "—"


def extract_pack_size(title: str, handle: str, variant_title: str, metafield_value: str = "") -> str:
    if metafield_value and metafield_value != "—":
        return metafield_value
    if variant_title and variant_title not in ("Default Title", "Default", "Single"):
        return variant_title
    for text in [title, handle]:
        m = re.search(r"(\d+\s*(?:ml|g|ML|G|mg|l|L|oz|pcs|pack|tab|capsules?|caps?))", text, re.I)
        if m:
            return m.group(1).strip().upper()
    return "—"


def sheet_name_from_title(title: str, max_len: int = 31) -> str:
    name = re.sub(r'[<>:"/\\|?*\[\]]', "", title)
    return name.strip()[:max_len]


# ─── Layer 2: Metafield Discovery & Mapping ──────────────────────────────────

def discover_metafield_schema(base_url: str, access_token: str) -> dict:
    """Query metafieldDefinitions to discover what custom fields exist on this store."""
    print("  Discovering store metafield schema...")

    query = """
    {
      metafieldDefinitions(ownerType: PRODUCT, first: 250) {
        edges {
          node {
            name
            namespace
            key
            type { name }
            description
          }
        }
      }
      productsCount { count }
    }
    """

    response = requests.post(
        f"{base_url}/admin/api/2026-01/graphql.json",
        headers={"X-Shopify-Access-Token": access_token, "Content-Type": "application/json"},
        json={"query": query},
        timeout=15,
    )

    data = response.json()
    if "errors" in data:
        print(f"  Warning: Schema discovery failed: {data['errors'][0].get('message', '')}")
        return {"definitions": [], "product_count": 0, "field_map": {}}

    defs = data.get("data", {}).get("metafieldDefinitions", {}).get("edges", [])
    count = data.get("data", {}).get("productsCount", {}).get("count", 0)

    definitions = []
    for d in defs:
        n = d["node"]
        definitions.append({
            "name": n.get("name", ""),
            "namespace": n.get("namespace", ""),
            "key": n.get("key", ""),
            "type": n.get("type", {}).get("name", ""),
            "description": n.get("description", ""),
        })

    # Build mapping: audit_field -> list of (namespace, key) pairs
    field_map = build_metafield_map(definitions)

    print(f"  Found {len(definitions)} metafield definitions, {count} total products")
    if field_map:
        mapped = [k for k, v in field_map.items() if v]
        print(f"  Auto-mapped metafields to: {', '.join(mapped)}")

    return {"definitions": definitions, "product_count": count, "field_map": field_map}


def build_metafield_map(definitions: list) -> dict:
    """Map metafield definitions to audit fields using keyword matching."""
    field_map = {k: [] for k in METAFIELD_MAPPING_RULES}

    for defn in definitions:
        # Build multiple search variants for flexible matching
        name_lower = defn['name'].lower()
        key_lower = defn['key'].lower()
        desc_lower = (defn.get('description') or '').lower()
        # Normalize: replace spaces and hyphens with underscores for matching
        key_normalized = key_lower.replace("-", "_").replace(" ", "_")
        name_normalized = name_lower.replace("-", "_").replace(" ", "_")
        search_variants = [key_lower, key_normalized, name_lower, name_normalized, desc_lower]

        ns_key = (defn["namespace"], defn["key"])

        for audit_field, keywords in METAFIELD_MAPPING_RULES.items():
            matched = False
            for kw in keywords:
                for variant in search_variants:
                    if kw in variant:
                        if ns_key not in field_map[audit_field]:
                            field_map[audit_field].append(ns_key)
                        matched = True
                        break
                if matched:
                    break

    return field_map


def find_metafield(metafields: dict, ns_key_pairs: list, fallback_keywords: list = None) -> str:
    """Find first matching metafield value by checking namespace.key substring matches, then keyword scan."""
    # Build a combined namespace+key string for each metafield for substring matching
    all_keys = list(metafields.keys())

    # First: try mapped keys from schema discovery using substring match on full ns.key string
    for ns, key in ns_key_pairs:
        if not key:
            continue
        # Try exact match first
        exact_key = f"{ns}.{key}" if ns else key
        if exact_key in metafields:
            val = metafields[exact_key]
            if val and val != "—":
                return val
        # Try substring match: if mapped key is 'how_it_works_description',
        # match 'how_it_works' in 'custom.how_it_works_description'
        search_str = (ns + "." + key).lower()
        for full_key, val in metafields.items():
            if not val or val == "—":
                continue
            if search_str in full_key.lower():
                return val

    # Second: if no mapped keys matched, scan ALL metafields for keyword matches
    if fallback_keywords:
        for full_key, val in metafields.items():
            if not val or val == "—":
                continue
            key_lower = full_key.lower()
            for kw in fallback_keywords:
                if kw in key_lower:
                    return val

    return ""


# ─── Review App Detection ────────────────────────────────────────────────────

def extract_rating(metafields: dict) -> tuple:
    """Detect which review app is installed and extract rating + count."""
    for ns, rating_key, count_key, parser in REVIEW_APP_PATTERNS:
        rating_raw = metafields.get(f"{ns}.{rating_key}", "")
        if not rating_raw:
            continue

        rating_val = ""
        count_val = ""

        if parser == "shopify_standard":
            # JSON format: {"scale_min":"1.0","scale_max":"5.0","value":"4.81"}
            try:
                rdata = json.loads(rating_raw) if isinstance(rating_raw, str) else rating_raw
                rating_val = rdata.get("value", "")
                count_raw = metafields.get(f"{ns}.{count_key}", "")
                count_val = str(count_raw) if count_raw else ""
            except (json.JSONDecodeError, AttributeError):
                continue

        elif parser == "judgeme_html":
            # HTML: data-average-rating='4.81' data-number-of-reviews='103'
            m_rating = re.search(r"data-average-rating=['\"]([0-9.]+)['\"]", rating_raw)
            m_count = re.search(r"data-number-of-reviews=['\"](\d+)['\"]", rating_raw)
            if m_rating:
                rating_val = m_rating.group(1)
            if m_count:
                count_val = m_count.group(1)

        elif parser == "amazon_reviews":
            # ReputonAmazonReviews::ProductReviews.{id} — JSON: {"name","reviewsNumber","rating","imageUrl","url"}
            try:
                rdata = json.loads(rating_raw) if isinstance(rating_raw, str) else rating_raw
                rating_val = str(rdata.get("rating", ""))
                count_val = str(rdata.get("reviewsNumber", ""))
                amazon_url = rdata.get("url", "")
                if rating_val and count_val:
                    return f"{rating_val}/5 ({count_val} Amazon reviews)", count_val
            except (json.JSONDecodeError, AttributeError):
                continue

        elif parser == "direct":
            rating_val = str(rating_raw).strip()
            if count_key:
                count_val = str(metafields.get(f"{ns}.{count_key}", "")).strip()

        if rating_val:
            try:
                rv = float(rating_val)
                rating_str = f"{rv:.2f} / 5.0"
                if count_val and count_val.isdigit() and int(count_val) > 0:
                    rating_str += f" ({count_val} reviews)"
                return rating_str, count_val or "—"
            except ValueError:
                continue

    return "—", "—"


# ─── Layer 3: Description Parsing ────────────────────────────────────────────

def parse_description_sections(html: str) -> dict:
    """Split description HTML on heading tags and classify each section."""
    if not html:
        return {}

    sections = {}
    # Split on h1-h6 tags
    parts = re.split(r"<h[1-6][^>]*>(.*?)</h[1-6]>", html, flags=re.I | re.DOTALL)

    # parts alternates: [before_first_heading, heading1, content1, heading2, content2, ...]
    for i in range(1, len(parts) - 1, 2):
        heading_text = clean_html(parts[i]).lower().strip()
        content = clean_html(parts[i + 1]).strip()
        if not content:
            continue

        for category, keywords in HEADING_CLASSIFICATION.items():
            if any(kw in heading_text for kw in keywords):
                sections[category] = content
                break

    return sections


def extract_molecules_from_text(text: str) -> str:
    """Extract known molecules/ingredients from text."""
    if not text:
        return "—"
    found = []
    text_lower = text.lower()
    for mol in MOLECULES:
        if mol in text_lower:
            display = mol.title()
            if mol == "vitamin c":
                display = "Vitamin C"
            elif mol == "spf":
                display = "SPF"
            elif mol == "aha":
                display = "AHA"
            elif mol == "bha":
                display = "BHA"
            found.append(display)

    seen = set()
    unique = [x for x in found if not (x.lower() in seen or seen.add(x.lower()))]
    return ", ".join(unique[:10]) if unique else "—"


def extract_skin_types_from_text(text: str) -> str:
    """Extract skin type mentions from text."""
    if not text:
        return "—"
    text_lower = text.lower()
    found = [label for kw, label in SKIN_TYPE_KEYWORDS.items() if kw in text_lower]
    return ", ".join(sorted(set(found))) if found else "—"


def extract_certs_from_text(text: str) -> str:
    """Extract certification keywords from text."""
    if not text:
        return "—"
    text_lower = text.lower()
    found = []
    for kw in CERT_KEYWORDS:
        if kw in text_lower:
            found.append(kw.title())
    # Normalize common certs
    cleaned = []
    for c in found:
        if "Clinically Proven" in c:
            cleaned.append("Clinically Proven")
        elif "Dermatologist" in c and "Formulated" in c:
            cleaned.append("Dermatologist Formulated")
        elif "Dermatologist" in c and "Tested" in c:
            cleaned.append("Dermatologist Tested")
        elif "Dermatologist" in c and "Recommended" in c:
            cleaned.append("Dermatologist Recommended")
        elif "Fda" in c:
            cleaned.append("FDA Approved")
        elif "Iso" in c:
            cleaned.append("ISO Certified")
        elif "Made Safe" in c:
            cleaned.append("Made Safe Certified")
        elif "Gmp" in c:
            cleaned.append("GMP Certified")
        else:
            cleaned.append(c)
    return ", ".join(sorted(set(cleaned))) if cleaned else "—"


def classify_tags(tags: list) -> dict:
    """Classify tags into buckets: skin_types, concerns, ingredients, categories."""
    result = {"skin_types": [], "concerns": [], "ingredients": [], "categories": [], "other": []}
    if not tags:
        return result

    concern_keywords = [
        "acne", "aging", "anti-aging", "brightening", "dark-spot", "pigmentation",
        "hyperpigmentation", "hydration", "wrinkle", "fine-line", "pore", "redness",
        "dullness", "uneven", "melasma", "tan", "sun-damage", "stretch",
    ]
    category_keywords = [
        "serum", "moisturizer", "cleanser", "toner", "sunscreen", "spf", "mask",
        "eye-cream", "exfoliant", "face-wash", "body-lotion", "lip-balm", "face-oil",
        "mist", "cream", "gel", "lotion", "facial kit",
    ]
    marketing_keywords = [
        "best-seller", "bestseller", "new-arrival", "sale", "badge_", "featured",
        "trending", "buy1get1", "bogo", "limited", "april-sell", "sale-timer",
    ]

    for tag in tags:
        tag_lower = tag.lower().strip()
        if any(mk in tag_lower for mk in marketing_keywords):
            continue  # skip marketing tags

        matched = False
        for kw, label in SKIN_TYPE_KEYWORDS.items():
            if kw in tag_lower:
                result["skin_types"].append(label)
                matched = True
                break
        if matched:
            continue

        for ck in concern_keywords:
            if ck in tag_lower:
                result["concerns"].append(tag.strip().title())
                matched = True
                break
        if matched:
            continue

        for mol in MOLECULES:
            if mol in tag_lower:
                result["ingredients"].append(mol.title())
                matched = True
                break
        if matched:
            continue

        for cat in category_keywords:
            if cat in tag_lower:
                result["categories"].append(tag.strip().title())
                matched = True
                break
        if matched:
            continue

        result["other"].append(tag.strip())

    # Deduplicate
    for k in result:
        result[k] = list(dict.fromkeys(result[k]))

    return result


# ─── Metafield Value Parsers ─────────────────────────────────────────────────

def parse_metafield_value(value: str, mf_type: str = "") -> str:
    """Parse a metafield value into a human-readable string."""
    if not value:
        return ""

    # JSON list -> comma-separated
    if value.startswith("["):
        try:
            items = json.loads(value)
            # Filter out GID references (metaobject refs we can't resolve)
            if items and isinstance(items[0], str) and items[0].startswith("gid://"):
                return f"[{len(items)} metaobject references]"
            return ", ".join(str(i).strip() for i in items if i)
        except json.JSONDecodeError:
            pass

    # JSON object (rating type)
    if value.startswith("{"):
        try:
            obj = json.loads(value)
            if "value" in obj and "scale_max" in obj:
                return f"{obj['value']} / {obj['scale_max']}"
            return json.dumps(obj)
        except json.JSONDecodeError:
            pass

    # HTML content -> clean
    if "<" in value and ">" in value:
        return clean_html(value)

    return value.strip()


# ─── Shopify API ─────────────────────────────────────────────────────────────

def get_shopify_token(base_url: str, client_id: str, client_secret: str) -> str:
    """Get access token via client_credentials grant."""
    print(f"  Getting OAuth token for {base_url}...")
    response = requests.post(
        f"{base_url}/admin/oauth/access_token",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=15,
    )
    if response.status_code != 200:
        raise Exception(f"OAuth failed ({response.status_code}): {response.text[:300]}")

    data = response.json()
    expires = data.get("expires_in", 0)
    print(f"  Token received (expires in {expires // 3600}h {(expires % 3600) // 60}m)")
    return data["access_token"]


def fetch_products_graphql(base_url: str, access_token: str) -> list:
    """Fetch ALL products with cursor-based pagination."""
    print("  Fetching products via GraphQL (paginated)...")

    query = """
    query Products($cursor: String) {
      products(first: 50, after: $cursor) {
        edges {
          node {
            id
            title
            handle
            vendor
            productType
            tags
            status
            descriptionHtml
            category { name fullName }
            seo { title description }
            onlineStoreUrl
            totalInventory
            createdAt
            updatedAt
            publishedAt
            collections(first: 20) {
              edges { node { title handle } }
            }
            variants(first: 20) {
              edges {
                node {
                  title
                  price
                  compareAtPrice
                  sku
                  barcode
                  availableForSale
                  inventoryQuantity
                  selectedOptions { name value }
                }
              }
            }
            images(first: 5) {
              edges { node { url altText } }
            }
            metafields(first: 50) {
              edges {
                node {
                  namespace
                  key
                  value
                  type
                }
              }
            }
          }
        }
        pageInfo {
          hasNextPage
          endCursor
        }
      }
    }
    """

    all_products = []
    cursor = None
    page = 0

    while True:
        page += 1
        variables = {"cursor": cursor} if cursor else {}

        response = requests.post(
            f"{base_url}/admin/api/2026-01/graphql.json",
            headers={
                "X-Shopify-Access-Token": access_token,
                "Content-Type": "application/json",
            },
            json={"query": query, "variables": variables},
            timeout=30,
        )

        if response.status_code != 200:
            raise Exception(f"GraphQL failed ({response.status_code}): {response.text[:300]}")

        data = response.json()
        if "errors" in data:
            raise Exception(f"GraphQL errors: {json.dumps(data['errors'][:2])}")

        edges = data["data"]["products"]["edges"]
        page_info = data["data"]["products"]["pageInfo"]

        all_products.extend(edges)
        print(f"    Page {page}: {len(edges)} products (total: {len(all_products)})")

        if not page_info["hasNextPage"]:
            break

        cursor = page_info["endCursor"]

    print(f"  Got {len(all_products)} products total")
    return all_products


def fetch_products_public_json(url: str) -> list:
    """Fetch ALL products from public /products.json with pagination."""
    print(f"  Fetching from public endpoint (paginated)...")
    all_products = []
    page = 1

    while True:
        sep = "&" if "?" in url else "?"
        page_url = f"{url}{sep}limit=250&page={page}"
        resp = requests.get(page_url, timeout=15, headers={"User-Agent": "BrandAuditExtractor/1.0"})
        resp.raise_for_status()
        products = resp.json().get("products", [])

        if not products:
            break

        all_products.extend(products)
        print(f"    Page {page}: {len(products)} products (total: {len(all_products)})")

        if len(products) < 250:
            break
        page += 1

    print(f"  Got {len(all_products)} products total")
    return all_products


# ─── GA4 Extraction ──────────────────────────────────────────────────────────

def fetch_ga4(property_id: str, credentials_path: str) -> dict:
    """Fetch all GA4 data using the Google Analytics Data API."""
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = credentials_path

    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import RunReportRequest

    client = BetaAnalyticsDataClient()
    prop = f"properties/{property_id}"
    data = {}

    def run(dim_names, met_names, start="30daysAgo", end="today"):
        req = RunReportRequest(
            property=prop,
            date_ranges=[{"start_date": start, "end_date": end}],
            dimensions=[{"name": n} for n in dim_names],
            metrics=[{"name": n} for n in met_names],
        )
        resp = client.run_report(req)
        return resp.rows or []

    rows = run(["date"], ["sessions", "totalUsers", "screenPageViews", "bounceRate", "engagementRate", "activeUsers"])
    data["summary"] = {
        "total_sessions": sum(int(r.metric_values[0].value) for r in rows),
        "total_users": sum(int(r.metric_values[1].value) for r in rows),
        "total_pageviews": sum(int(r.metric_values[2].value) for r in rows),
        "avg_bounce_rate": sum(float(r.metric_values[3].value) for r in rows) / max(len(rows), 1),
        "avg_engagement_rate": sum(float(r.metric_values[4].value) for r in rows) / max(len(rows), 1),
        "date_range": "Last 30 days",
    }

    rows = run(["sessionDefaultChannelGroup"], ["sessions", "totalUsers", "bounceRate", "engagementRate"])
    data["channels"] = sorted([{"channel": r.dimension_values[0].value, "sessions": int(r.metric_values[0].value), "users": int(r.metric_values[1].value), "bounce_rate": float(r.metric_values[2].value), "engagement_rate": float(r.metric_values[3].value)} for r in rows], key=lambda x: x["sessions"], reverse=True)

    rows = run(["pagePath"], ["screenPageViews", "sessions", "engagementRate", "averageSessionDuration"])
    data["top_pages"] = sorted([{"page": r.dimension_values[0].value, "views": int(r.metric_values[0].value), "sessions": int(r.metric_values[1].value), "engagement_rate": float(r.metric_values[2].value)} for r in rows], key=lambda x: x["views"], reverse=True)[:20]

    rows = run(["userAgeBracket"], ["totalUsers"])
    data["age_groups"] = [{"age_group": r.dimension_values[0].value, "users": int(r.metric_values[0].value)} for r in rows if r.dimension_values[0].value]

    rows = run(["userGender"], ["totalUsers"])
    data["gender"] = [{"gender": r.dimension_values[0].value, "users": int(r.metric_values[0].value)} for r in rows if r.dimension_values[0].value]

    rows = run(["city"], ["sessions"])
    data["top_cities"] = sorted([{"city": r.dimension_values[0].value, "sessions": int(r.metric_values[0].value)} for r in rows if r.dimension_values[0].value], key=lambda x: x["sessions"], reverse=True)[:15]

    rows = run(["country"], ["sessions"])
    data["top_countries"] = sorted([{"country": r.dimension_values[0].value, "sessions": int(r.metric_values[0].value)} for r in rows if r.dimension_values[0].value], key=lambda x: x["sessions"], reverse=True)[:10]

    rows = run(["deviceCategory"], ["sessions"])
    data["devices"] = sorted([{"device": r.dimension_values[0].value.title(), "sessions": int(r.metric_values[0].value)} for r in rows if r.dimension_values[0].value], key=lambda x: x["sessions"], reverse=True)

    rows = run(["date"], ["sessions", "totalUsers"])
    data["weekly_trend"] = [{"date": r.dimension_values[0].value, "sessions": int(r.metric_values[0].value), "users": int(r.metric_values[1].value)} for r in rows]

    return data


# ─── GSC Extraction ──────────────────────────────────────────────────────────

def fetch_gsc(site_url: str, credentials_path: str) -> dict:
    from googleapiclient.discovery import build
    from google.oauth2 import service_account
    from datetime import timedelta

    credentials = service_account.Credentials.from_service_account_file(
        credentials_path, scopes=["https://www.googleapis.com/auth/webmasters.readonly"],
    )
    service = build("webmasters", "v3", credentials=credentials, cache_discovery=False)
    data = {}
    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")

    def query_gsc(dimensions, row_limit=1000):
        body = {"startDate": start_date, "endDate": end_date, "dimensions": dimensions, "rowLimit": row_limit}
        result = service.searchanalytics().query(siteUrl=site_url, body=body).execute()
        return result.get("rows", [])

    rows_all = query_gsc(dimensions=[], row_limit=1)
    if rows_all:
        r = rows_all[0]
        data["summary"] = {"total_clicks": int(r.get("clicks", 0)), "total_impressions": int(r.get("impressions", 0)), "avg_ctr": float(r.get("ctr", 0)), "avg_position": float(r.get("position", 0)), "date_range": f"{start_date} to {end_date}"}
    else:
        data["summary"] = {"total_clicks": 0, "total_impressions": 0, "avg_ctr": 0, "avg_position": 0, "date_range": f"{start_date} to {end_date}"}

    rows = query_gsc(["query"], 100)
    data["top_queries"] = sorted([{"query": r["keys"][0], "clicks": int(r.get("clicks", 0)), "impressions": int(r.get("impressions", 0)), "ctr": float(r.get("ctr", 0)), "position": float(r.get("position", 0))} for r in rows], key=lambda x: x["impressions"], reverse=True)

    rows = query_gsc(["page"], 50)
    data["top_pages"] = sorted([{"page": r["keys"][0], "clicks": int(r.get("clicks", 0)), "impressions": int(r.get("impressions", 0)), "ctr": float(r.get("ctr", 0)), "position": float(r.get("position", 0))} for r in rows], key=lambda x: x["impressions"], reverse=True)

    rows = query_gsc(["country"], 20)
    data["countries"] = sorted([{"country": r["keys"][0], "clicks": int(r.get("clicks", 0)), "impressions": int(r.get("impressions", 0)), "ctr": float(r.get("ctr", 0)), "position": float(r.get("position", 0))} for r in rows], key=lambda x: x["impressions"], reverse=True)

    rows = query_gsc(["device"], 10)
    data["devices"] = sorted([{"device": r["keys"][0], "clicks": int(r.get("clicks", 0)), "impressions": int(r.get("impressions", 0)), "ctr": float(r.get("ctr", 0)), "position": float(r.get("position", 0))} for r in rows], key=lambda x: x["impressions"], reverse=True)

    rows = query_gsc(["query"], 1000)
    data["zero_clicks"] = sorted([{"query": r["keys"][0], "clicks": 0, "impressions": int(r.get("impressions", 0)), "ctr": float(r.get("ctr", 0)), "position": float(r.get("position", 0))} for r in rows if int(r.get("impressions", 0)) >= 100 and int(r.get("clicks", 0)) == 0], key=lambda x: x["impressions"], reverse=True)

    return data


# ─── Universal Product Data Extraction ───────────────────────────────────────

def extract_product_universal(p: dict, base_url: str, source: str, field_map: dict) -> dict:
    """
    3-layer extraction:
      Layer 1: Guaranteed Shopify fields
      Layer 2: Auto-mapped metafields
      Layer 3: Fallback parsing from description/tags
    """

    # ── Parse raw data based on source ──────────────────────────────────────
    if source == "graphql":
        node = p["node"]
        title = node.get("title", "")
        handle = node.get("handle", "")
        vendor = node.get("vendor", "")
        product_type = node.get("productType", "") or "—"
        tags = node.get("tags", [])
        status = node.get("status", "")
        desc_html = node.get("descriptionHtml", "")

        # Category
        category_obj = node.get("category")
        category = category_obj.get("fullName", "") if category_obj else ""

        # SEO
        seo = node.get("seo", {})
        seo_title = seo.get("title", "")
        seo_desc = seo.get("description", "")

        # URL
        pdp_url = node.get("onlineStoreUrl") or f"{base_url}/products/{handle}"

        # Collections
        collections = [e["node"]["title"] for e in node.get("collections", {}).get("edges", [])]

        # Variants
        variant_edges = node.get("variants", {}).get("edges", [])
        if variant_edges:
            variant = variant_edges[0]["node"]
            price = variant.get("price", "")
            mrp = variant.get("compareAtPrice") or ""
            variant_title = variant.get("title", "")
            sku = variant.get("sku", "")
        else:
            price, mrp, variant_title, sku = "", "", "", ""

        if not price:
            price = node.get("priceRange", {}).get("minVariantPrice", {}).get("amount", "")

        # Inventory
        total_inventory = node.get("totalInventory", 0)

        # Published date
        published_at = node.get("publishedAt", "")

        # ALL metafields (no namespace filter!)
        metafields = {}
        for mf_edge in node.get("metafields", {}).get("edges", []):
            mf = mf_edge.get("node", {})
            if mf:
                ns = mf.get("namespace", "")
                key = mf.get("key", "")
                val = mf.get("value", "")
                mf_type = mf.get("type", "")
                if ns and key and val:
                    metafields[f"{ns}.{key}"] = val

        # Google Shopping labels
        google_labels = []
        for k, v in metafields.items():
            if k.startswith("mm-google-shopping.custom_label"):
                if v and v not in google_labels:
                    google_labels.append(str(v))

        # Before/after images
        before_img = metafields.get("custom.before_after_section_before", "")
        after_img = metafields.get("custom.after_product_use_result", "") or metafields.get("custom.after_product_use_result", "")
        if not after_img:
            after_img = metafields.get("custom.before_after_section_after", "")

        # Amazon reviews
        amazon_reviews = "—"
        for k, v in metafields.items():
            if k.startswith("reputon") and "reviewsNumber" in v:
                try:
                    rdata = json.loads(v) if isinstance(v, str) else v
                    count = rdata.get("reviewsNumber", "")
                    rating_val = rdata.get("rating", "")
                    url = rdata.get("url", "")
                    if count:
                        amazon_reviews = f"{rating_val}/5 ({count} Amazon reviews) — {url}"
                except (json.JSONDecodeError, AttributeError):
                    pass

        # Doctor profiles (metaobject references)
        doctor_profiles_raw = find_metafield(metafields, field_map.get("doctor_involvement", []),
                                              fallback_keywords=["doctor_profile", "doctor_profiles"])
        doctor_profiles = parse_metafield_value(doctor_profiles_raw) if doctor_profiles_raw else "—"

        # Related products
        related_prods = []
        for k, v in metafields.items():
            if "related_product" in k.lower() or "related_products" in k.lower():
                if v.startswith("["):
                    try:
                        refs = json.loads(v)
                        if refs and isinstance(refs[0], str):
                            related_prods.append(f"[{len(refs)} product references]")
                    except json.JSONDecodeError:
                        pass
        related_prods_str = ", ".join(related_prods) if related_prods else "—"

        # Product images
        img_edges = node.get("images", {}).get("edges", [])
        img_urls = [e["node"]["url"] for e in img_edges[:5]]
        primary_img = img_urls[0] if img_urls else "—"

        # EAN code
        ean = metafields.get("custom.ean_code", "") or "—"

        # Shelf life / storage
        shelf_raw = find_metafield(metafields, field_map.get("shelf_life", []),
                                   fallback_keywords=["shelf_life", "dermatologist_recommended", "storage"])
        shelf_parsed = parse_metafield_value(shelf_raw) if shelf_raw else "—"

        # Texture experience
        texture_raw = find_metafield(metafields, field_map.get("texture_experience", []),
                                     fallback_keywords=["texture_experience"])
        texture_parsed = parse_metafield_value(texture_raw) if texture_raw else "—"

        # Who shouldn't use
        who_shouldnt_raw = find_metafield(metafields, field_map.get("who_shouldnt", []),
                                           fallback_keywords=["shouldn_t", "should_not"])
        who_shouldnt_parsed = parse_metafield_value(who_shouldnt_raw) if who_shouldnt_raw else "—"

        # Results / why you'll love it
        results_raw = find_metafield(metafields, field_map.get("results_why_love", []),
                                      fallback_keywords=["results", "why_choose"])
        results_parsed = parse_metafield_value(results_raw) if results_raw else "—"

        # Detailed description
        detailed_desc_raw = find_metafield(metafields, field_map.get("detailed_description", []),
                                            fallback_keywords=["detailed_product_description"])
        detailed_desc_parsed = parse_metafield_value(detailed_desc_raw) if detailed_desc_raw else "—"

        # Short description
        short_desc_raw = find_metafield(metafields, field_map.get("short_desc_field", []),
                                         fallback_keywords=["short_description", "product_short_description"])
        short_desc_parsed = parse_metafield_value(short_desc_raw) if short_desc_raw else "—"

        # Product USPs
        usps_raw = find_metafield(metafields, field_map.get("usps", []),
                                   fallback_keywords=["usps", "product_usps"])
        usps_parsed = parse_metafield_value(usps_raw) if usps_raw else "—"

        # FAQ content
        faq_raw = find_metafield(metafields, field_map.get("faq_content", []),
                                  fallback_keywords=["product_faq", "faq", "faqs"])
        faq_parsed = parse_metafield_value(faq_raw) if faq_raw else "—"

        # Breadcrumb
        breadcrumb_raw = find_metafield(metafields, field_map.get("breadcrumb", []),
                                         fallback_keywords=["breadcrumb", "custom_breadcrumb"])
        breadcrumb_parsed = clean_html(parse_metafield_value(breadcrumb_raw)) if breadcrumb_raw else "—"

    else:  # public JSON
        title = p.get("title", "")
        handle = p.get("handle", "")
        vendor = p.get("vendor", "")
        product_type = p.get("product_type", "") or "—"
        tags = p.get("tags", "")
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(",") if t.strip()]
        status = "ACTIVE"
        desc_html = p.get("body_html", "")
        category = ""
        seo_title = ""
        seo_desc = ""
        pdp_url = f"{base_url}/products/{handle}"
        collections = []
        variant = (p.get("variants") or [{}])[0] or {}
        price = variant.get("price", "")
        mrp = variant.get("compare_at_price") or ""
        variant_title = variant.get("title", "")
        sku = variant.get("sku", "")
        total_inventory = 0
        published_at = ""
        metafields = {}
        google_labels = []
        before_img = after_img = amazon_reviews = doctor_profiles = related_prods_str = "—"
        img_urls = []
        primary_img = "—"
        ean = sku or "—"
        shelf_parsed = texture_parsed = who_shouldnt_parsed = "—"
        results_parsed = detailed_desc_parsed = short_desc_parsed = "—"
        usps_parsed = faq_parsed = breadcrumb_parsed = "—"

    raw_desc = clean_html(desc_html)

    # ── Layer 1: Guaranteed fields ──────────────────────────────────────────
    result = {
        "Product Name": title,
        "PDP URL (Clean)": pdp_url,
        "Product Subtitle": "—",
        "Selling Price (INR)": f"Rs.{float(price):.0f}" if price else "—",
        "MRP / Compare-at (INR)": f"Rs.{float(mrp):.0f}" if mrp else "—",
        "Discount %": get_discount(price, mrp),
        "SKU / EAN": ean if ean else (sku or "—"),
        "Product Type": product_type,
        "Pack Size": "—",
        "Total Inventory": str(total_inventory) if total_inventory else "—",
        "Category (Taxonomy)": category or "—",
        "Rating": "—",
        "Review Count": "—",
        "Amazon Reviews": amazon_reviews,
        "Hero Molecules": "—",
        "Full Ingredient List": "—",
        "Clinical Claims": "—",
        "Clinical Method": "—",
        "How It Works": "—",
        "Target Skin Types": "—",
        "Who It's For": "—",
        "Who It's NOT For": "—",
        "Certifications": "—",
        "Dr. Involvement": "—",
        "Doctor Profiles": doctor_profiles,
        "Results (Why You'll Love It)": "—",
        "Texture / Experience": "—",
        "How To Use": "—",
        "Shelf Life & Storage": shelf_parsed,
        "Before/After Images": f"BEFORE: {before_img[:80]}... AFTER: {after_img[:80]}..." if before_img or after_img else "—",
        "Collections": ", ".join(collections[:8]) if collections else "—",
        "Related Products": related_prods_str,
        "SEO Title": seo_title,
        "SEO Description": seo_desc[:200] if seo_desc else "—",
        "Breadcrumb": breadcrumb_parsed,
        "Google Shopping Labels": ", ".join(google_labels) if google_labels else "—",
        "Short Description": short_desc_parsed,
        "Product USPs": usps_parsed[:300] if usps_parsed else "—",
        "FAQ Content": faq_parsed[:300] if faq_parsed else "—",
        "Published Date": published_at[:10] if published_at else "—",
        "Price Positioning": "—",
        "_status": status,
        "_tags": tags,
        "_source": source,
        "_vendor": vendor,
        "_sku": sku,
        "_primary_image": primary_img,
    }

    # ── Layer 2: Metafield-based extraction ─────────────────────────────────

    # Rating (multi-app detection)
    rating_str, review_count = extract_rating(metafields)
    result["Rating"] = rating_str
    result["Review Count"] = review_count

    # Hero Molecules / Key Ingredients
    hero_raw = find_metafield(metafields, field_map.get("hero_molecules", []),
                              fallback_keywords=["product_ingredient", "key_ingredient", "hero", "product_ingredients"])
    hero_parsed = parse_metafield_value(hero_raw) if hero_raw else ""
    result["Hero Molecules"] = hero_parsed or "—"

    # Full Ingredient List (broadest ingredient search)
    ingr_raw = find_metafield(metafields, field_map.get("full_ingredient_list", []),
                              fallback_keywords=["ingredient_list", "active_ingredient", "ingredient"])
    ingr_parsed = parse_metafield_value(ingr_raw) if ingr_raw else ""
    result["Full Ingredient List"] = ingr_parsed or "—"

    # Target Skin Types
    skin_raw = find_metafield(metafields, field_map.get("target_skin_types", []),
                              fallback_keywords=["skin_type", "who_is_it_for", "who_should", "who_should_use_it"])
    skin_parsed = parse_metafield_value(skin_raw) if skin_raw else ""
    result["Target Skin Types"] = skin_parsed or "—"

    # Who It's For
    who_raw = find_metafield(metafields, field_map.get("who_its_for", []),
                             fallback_keywords=["who_is_it_for", "suitable", "ideal_for", "who_is_it_for_items"])
    who_parsed = parse_metafield_value(who_raw) if who_raw else ""
    result["Who It's For"] = who_parsed or "—"

    # Clinical Claims — try results metafield first, clean HTML, then fallback to description HTML
    # If custom.clinically_proven_results returns metaobject refs, treat as "not found"
    clinical_raw = find_metafield(metafields, field_map.get("clinical_claims", []),
                                  fallback_keywords=["results", "why_love", "why_choose_skinq", "clinical", "proven"])
    clinical_parsed = parse_metafield_value(clinical_raw) if clinical_raw else ""
    # If metaobject refs, treat as empty — will be enriched from description HTML
    if clinical_parsed and "metaobject" in clinical_parsed.lower():
        clinical_parsed = ""
    # Clean HTML from results field (my_fields.results contains HTML tags)
    if clinical_parsed and clinical_parsed != "—" and "<" in clinical_parsed:
        clinical_parsed = clean_html(clinical_parsed)
    result["Clinical Claims"] = clinical_parsed or "—"

    # Certifications
    cert_raw = find_metafield(metafields, field_map.get("certifications", []),
                              fallback_keywords=["certification", "certified", "why_choose_skinq"])
    cert_parsed = parse_metafield_value(cert_raw) if cert_raw else ""
    if cert_parsed and cert_parsed != "—" and "<" in cert_parsed:
        cert_parsed = clean_html(cert_parsed)
    result["Certifications"] = cert_parsed or "—"

    # Doctor Involvement
    doc_raw = find_metafield(metafields, field_map.get("doctor_involvement", []),
                             fallback_keywords=["doctor", "dermatologist", "formulator"])
    doc_parsed = parse_metafield_value(doc_raw) if doc_raw else ""
    if doc_parsed and doc_parsed != "—":
        result["Dr. Involvement"] = doc_parsed

    # Subtitle
    sub_raw = find_metafield(metafields, field_map.get("subtitle", []),
                             fallback_keywords=["subtitle", "sub_title", "product_sub_title"])
    sub_parsed = parse_metafield_value(sub_raw) if sub_raw else ""
    result["Product Subtitle"] = sub_parsed or "—"

    # Pack Size
    size_raw = find_metafield(metafields, field_map.get("pack_size", []),
                              fallback_keywords=["pack_size", "net_weight", "volume"])
    size_parsed = parse_metafield_value(size_raw) if size_raw else ""
    result["Pack Size"] = extract_pack_size(title, handle, variant_title, size_parsed)

    # How It Works
    how_it_works_raw = find_metafield(metafields, field_map.get("how_it_works", []),
                                       fallback_keywords=["how_it_works_description", "how_it_works"])
    how_it_works_parsed = parse_metafield_value(how_it_works_raw) if how_it_works_raw else ""
    result["How It Works"] = how_it_works_parsed[:300] if how_it_works_parsed else "—"

    # How To Use
    how_to_use_raw = find_metafield(metafields, field_map.get("how_to_use", []),
                                     fallback_keywords=["how_to_use_", "how_to_use"])
    how_to_use_parsed = parse_metafield_value(how_to_use_raw) if how_to_use_raw else ""
    result["How To Use"] = how_to_use_parsed[:300] if how_to_use_parsed else "—"

    # Results / Why You'll Love It
    if results_parsed == "—" or not results_parsed:
        results_raw = find_metafield(metafields, field_map.get("results_why_love", []),
                                       fallback_keywords=["results"])
        results_parsed = parse_metafield_value(results_raw)[:300] if results_raw else "—"
    result["Results (Why You'll Love It)"] = results_parsed if results_parsed != "—" else "—"

    # ── Layer 3: Fallback enrichment from description + tags ────────────────
    desc_sections = parse_description_sections(desc_html)
    tag_buckets = classify_tags(tags)

    # Enrich Hero Molecules if empty or unresolvable metaobject refs
    if result["Hero Molecules"] == "—" or "metaobject" in result["Hero Molecules"]:
        if result["Full Ingredient List"] != "—" and "metaobject" not in result["Full Ingredient List"]:
            molecules_from_inci = extract_molecules_from_text(result["Full Ingredient List"])
            if molecules_from_inci != "—":
                result["Hero Molecules"] = molecules_from_inci
        if result["Hero Molecules"] == "—" or "metaobject" in result["Hero Molecules"]:
            if "INGREDIENTS" in desc_sections:
                result["Hero Molecules"] = extract_molecules_from_text(desc_sections["INGREDIENTS"])
        if result["Hero Molecules"] == "—" or "metaobject" in result["Hero Molecules"]:
            extracted = extract_molecules_from_text(raw_desc)
            if extracted != "—":
                result["Hero Molecules"] = extracted
        if result["Hero Molecules"] == "—" or "metaobject" in result["Hero Molecules"]:
            if tag_buckets["ingredients"]:
                result["Hero Molecules"] = ", ".join(tag_buckets["ingredients"][:8])

    # Enrich Target Skin Types if still empty
    if result["Target Skin Types"] == "—":
        if "SKIN_TYPE" in desc_sections:
            result["Target Skin Types"] = extract_skin_types_from_text(desc_sections["SKIN_TYPE"])
        if result["Target Skin Types"] == "—" and tag_buckets["skin_types"]:
            result["Target Skin Types"] = ", ".join(tag_buckets["skin_types"])
        if result["Target Skin Types"] == "—":
            result["Target Skin Types"] = extract_skin_types_from_text(raw_desc + " " + " ".join(tags))

    # Enrich Certifications if still empty
    if result["Certifications"] == "—":
        result["Certifications"] = extract_certs_from_text(raw_desc + " " + " ".join(tags))

    # Enrich Clinical Claims if still empty or unresolvable metaobject refs
    if result["Clinical Claims"] == "—" or "metaobject" in result["Clinical Claims"].lower():
        if "CLINICAL" in desc_sections:
            result["Clinical Claims"] = desc_sections["CLINICAL"][:500]
        elif raw_desc:
            # Try to find "Clinically Proven" or "Proven" sections in raw description
            clinical_match = re.search(
                r'(?:Clinically Proven|clinical study|clinical evaluation| proven to |proved to |Based on clinical).*',
                raw_desc, re.I | re.DOTALL
            )
            if clinical_match:
                result["Clinical Claims"] = clinical_match.group(0)[:500].strip()

    # Enrich Who It's For from tags
    if result["Who It's For"] == "—" and tag_buckets["concerns"]:
        result["Who It's For"] = ", ".join(tag_buckets["concerns"][:6])

    # Clinical Method from description sections
    if result["Clinical Method"] == "—" and "CLINICAL" in desc_sections:
        result["Clinical Method"] = desc_sections["CLINICAL"][:200]

    return result


# ─── Excel Builder ───────────────────────────────────────────────────────────

def get_field_color(field_name: str) -> str:
    manual_fields = {"Price Positioning", "Dr. Involvement", "Clinical Method"}
    if field_name in manual_fields:
        return "FFF2CC"  # Orange/yellow for manual fill
    return "FFFFFF"


def build_excel(products: list, brand_name: str, base_url: str,
                today: str, output_path: str,
                ga4_data: dict = None, ga4_property: str = "",
                gsc_data: dict = None, gsc_site: str = "") -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary Sheet (first) ─────────────────────────────────────────────
    ws = wb.create_sheet(title="Audit Summary")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Brand Audit Summary")
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Extracted: {today} | Products: {len(products)} | Source: Shopify Admin API")
    ws.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True)

    # Summary headers
    summary_headers = ["#", "Product Name", "Status", "Price", "MRP", "Discount", "Type", "Rating", "Category"]
    for col, hdr in enumerate(summary_headers, 1):
        ws.cell(row=4, column=col, value=hdr)
        ws.cell(row=4, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=4, column=col).fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9D9D9")

    for i, p in enumerate(products):
        r = 5 + i
        for col, val in enumerate([
            i + 1, p.get("Product Name", ""), p.get("_status", ""),
            p.get("Selling Price (INR)", ""), p.get("MRP / Compare-at (INR)", ""),
            p.get("Discount %", ""), p.get("Product Type", ""),
            p.get("Rating", ""), p.get("Category (Taxonomy)", ""),
        ], 1):
            ws.cell(row=r, column=col, value=val)
        if i % 2 == 0:
            for col in range(1, len(summary_headers) + 1):
                ws.cell(row=r, column=col).fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="F2F2F2")

    widths = [5, 45, 10, 12, 12, 10, 20, 25, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # ── Product Sheets ────────────────────────────────────────────────────
    for idx, p in enumerate(products, 1):
        title = p.get("Product Name", "")
        sheet_name = sheet_name_from_title(title, 31)
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{idx}"

        ws = wb.create_sheet(title=sheet_name)
        source = p.get("_source", "public")

        ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Product {idx}: {title}")
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=12)

        source_line = "SOURCE: Shopify Admin API (GraphQL)" if source == "graphql" else f"SOURCE: {base_url}/products.json"
        ws.cell(row=2, column=1, value=f"{source_line} | Extracted: {today}")
        ws.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True)
        ws.cell(row=3, column=1, value="WARNING: Yellow = manual fill required.")

        ws.cell(row=5, column=1, value="Field")
        ws.cell(row=5, column=2, value="Value")
        ws.cell(row=5, column=3, value="Audit Input Mapping")
        ws.cell(row=5, column=4, value="Action")
        for col in range(1, 5):
            ws.cell(row=5, column=col).font = openpyxl.styles.Font(bold=True)
            ws.cell(row=5, column=col).fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9D9D9")

        for row_idx, (field, mapping, action) in enumerate(AUDIT_FIELDS, start=6):
            value = p.get(field, "—")
            ws.cell(row=row_idx, column=1, value=field)
            ws.cell(row=row_idx, column=2, value=str(value)[:500] if value else "—")
            ws.cell(row=row_idx, column=3, value=mapping)
            ws.cell(row=row_idx, column=4, value=action)
            color = get_field_color(field)
            ws.cell(row=row_idx, column=2).fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=color)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 40

    # ── GA4 Sheets ────────────────────────────────────────────────────────
    if ga4_data:
        _build_ga4_sheets(wb, ga4_data, brand_name, ga4_property, today)

    # ── GSC Sheets ────────────────────────────────────────────────────────
    if gsc_data:
        _build_gsc_sheets(wb, gsc_data, brand_name, gsc_site, today)

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


def _build_ga4_sheets(wb, ga4_data, brand_name, ga4_property, today):
    s = ga4_data.get("summary", {})
    hdr_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9D9D9")
    alt_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="F2F2F2")

    # Summary
    ws = wb.create_sheet(title="GA4 Summary")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- GA4 Traffic Summary").font = openpyxl.styles.Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Property: {ga4_property} | Extracted: {today}").font = openpyxl.styles.Font(italic=True)
    for col, hdr in enumerate(["METRIC", "VALUE"], 1):
        ws.cell(row=4, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=4, column=col).fill = hdr_fill
    for i, (k, v) in enumerate([("Date Range", s.get("date_range", "—")), ("Total Sessions", f"{s.get('total_sessions', 0):,}"), ("Total Users", f"{s.get('total_users', 0):,}"), ("Total Pageviews", f"{s.get('total_pageviews', 0):,}"), ("Avg Bounce Rate", f"{s.get('avg_bounce_rate', 0):.1f}%"), ("Avg Engagement Rate", f"{s.get('avg_engagement_rate', 0):.1f}%")]):
        ws.cell(row=5 + i, column=1, value=k)
        ws.cell(row=5 + i, column=2, value=v)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # Channels
    ws = wb.create_sheet(title="GA4 Channels")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Traffic by Channel (Last 30 Days)").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["Channel", "Sessions", "Users", "Bounce Rate", "Engagement Rate", "% of Total"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=3, column=col).fill = hdr_fill
    total_sess = s.get("total_sessions", 1)
    for i, ch in enumerate(ga4_data.get("channels", [])):
        r = 4 + i
        pct = ch["sessions"] / total_sess * 100
        for col, val in enumerate([ch["channel"], ch["sessions"], ch["users"], f"{ch['bounce_rate']:.1f}%", f"{ch['engagement_rate']:.1f}%", f"{pct:.1f}%"], 1):
            ws.cell(row=r, column=col, value=val)
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = alt_fill

    # Top Pages
    ws = wb.create_sheet(title="GA4 Top Pages")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Top Pages (Last 30 Days)").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["Page URL", "Views", "Sessions", "Engagement Rate", "Notes"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=3, column=col).fill = hdr_fill
    for i, pg in enumerate(ga4_data.get("top_pages", [])):
        r = 4 + i
        note = "PDP" if "/products/" in pg["page"] else ("Collection" if "/collections/" in pg["page"] else "")
        for col, val in enumerate([pg["page"], pg["views"], pg["sessions"], f"{pg['engagement_rate']:.1f}%", note], 1):
            ws.cell(row=r, column=col, value=val)

    # Demographics
    ws = wb.create_sheet(title="GA4 Demographics")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Audience Demographics").font = openpyxl.styles.Font(bold=True, size=12)
    ws.cell(row=3, column=1, value="AGE GROUP").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=2, value="USERS").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=4, value="GENDER").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=5, value="USERS").font = openpyxl.styles.Font(bold=True)
    age_total = sum(a["users"] for a in ga4_data.get("age_groups", [])) or 1
    for i, ag in enumerate(ga4_data.get("age_groups", [])):
        ws.cell(row=4 + i, column=1, value=ag["age_group"])
        ws.cell(row=4 + i, column=2, value=f"{ag['users']} ({ag['users'] / age_total * 100:.0f}%)")
    gender_total = sum(g["users"] for g in ga4_data.get("gender", [])) or 1
    for i, g in enumerate(ga4_data.get("gender", [])):
        ws.cell(row=4 + i, column=4, value=g["gender"])
        ws.cell(row=4 + i, column=5, value=f"{g['users']} ({g['users'] / gender_total * 100:.0f}%)")

    # Locations
    ws = wb.create_sheet(title="GA4 Locations")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Top Locations").font = openpyxl.styles.Font(bold=True, size=12)
    ws.cell(row=3, column=1, value="CITY").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=2, value="SESSIONS").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=4, value="COUNTRY").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=5, value="SESSIONS").font = openpyxl.styles.Font(bold=True)
    for i, c in enumerate(ga4_data.get("top_cities", [])):
        ws.cell(row=4 + i, column=1, value=c["city"])
        ws.cell(row=4 + i, column=2, value=c["sessions"])
    for i, c in enumerate(ga4_data.get("top_countries", [])):
        ws.cell(row=4 + i, column=4, value=c["country"])
        ws.cell(row=4 + i, column=5, value=c["sessions"])

    # Devices
    ws = wb.create_sheet(title="GA4 Devices")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Device Breakdown").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["DEVICE", "SESSIONS", "% OF TOTAL"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
    dev_total = sum(d["sessions"] for d in ga4_data.get("devices", [])) or 1
    for i, d in enumerate(ga4_data.get("devices", [])):
        for col, val in enumerate([d["device"], d["sessions"], f"{d['sessions'] / dev_total * 100:.1f}%"], 1):
            ws.cell(row=4 + i, column=col, value=val)

    # Weekly Trend
    ws = wb.create_sheet(title="GA4 Weekly Trend")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Daily Sessions Trend").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["DATE", "SESSIONS", "USERS"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
    for i, d in enumerate(ga4_data.get("weekly_trend", [])):
        formatted = f"{d['date'][:4]}-{d['date'][4:6]}-{d['date'][6:]}"
        for col, val in enumerate([formatted, d["sessions"], d["users"]], 1):
            ws.cell(row=4 + i, column=col, value=val)


def _build_gsc_sheets(wb, gsc_data, brand_name, gsc_site, today):
    hdr_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9D9D9")
    alt_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="F2F2F2")
    s = gsc_data.get("summary", {})

    # Summary
    ws = wb.create_sheet(title="GSC Summary")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- GSC Search Performance").font = openpyxl.styles.Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Site: {gsc_site} | Extracted: {today}").font = openpyxl.styles.Font(italic=True)
    for col, hdr in enumerate(["METRIC", "VALUE"], 1):
        ws.cell(row=4, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=4, column=col).fill = hdr_fill
    for i, (k, v) in enumerate([("Date Range", s.get("date_range", "—")), ("Total Clicks", f"{s.get('total_clicks', 0):,}"), ("Total Impressions", f"{s.get('total_impressions', 0):,}"), ("Average CTR", f"{s.get('avg_ctr', 0) * 100:.2f}%"), ("Average Position", f"{s.get('avg_position', 0):.1f}")]):
        ws.cell(row=5 + i, column=1, value=k)
        ws.cell(row=5 + i, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25

    # Top Queries
    ws = wb.create_sheet(title="GSC Top Queries")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Top Queries (90 Days)").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["QUERY", "CLICKS", "IMPRESSIONS", "CTR", "POSITION"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=3, column=col).fill = hdr_fill
    for i, q in enumerate(gsc_data.get("top_queries", [])):
        r = 4 + i
        for col, val in enumerate([q["query"], q["clicks"], q["impressions"], f"{q['ctr'] * 100:.2f}%", f"{q['position']:.1f}"], 1):
            ws.cell(row=r, column=col, value=val)
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = alt_fill

    # Top Pages
    ws = wb.create_sheet(title="GSC Top Pages")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Top Pages (90 Days)").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["PAGE URL", "CLICKS", "IMPRESSIONS", "CTR", "POSITION"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=3, column=col).fill = hdr_fill
    for i, pg in enumerate(gsc_data.get("top_pages", [])):
        r = 4 + i
        for col, val in enumerate([pg["page"], pg["clicks"], pg["impressions"], f"{pg['ctr'] * 100:.2f}%", f"{pg['position']:.1f}"], 1):
            ws.cell(row=r, column=col, value=val)

    # Zero-Click Queries
    ws = wb.create_sheet(title="GSC Zero-Click Queries")
    ws.cell(row=1, column=1, value=f"{brand_name.upper()} -- Zero-Click Queries (Impressions >= 100, Clicks = 0)").font = openpyxl.styles.Font(bold=True, size=12)
    for col, hdr in enumerate(["QUERY", "IMPRESSIONS", "POSITION", "CTR"], 1):
        ws.cell(row=3, column=col, value=hdr).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=3, column=col).fill = hdr_fill
    for i, q in enumerate(gsc_data.get("zero_clicks", [])):
        for col, val in enumerate([q["query"], q["impressions"], f"{q['position']:.1f}", f"{q['ctr'] * 100:.2f}%"], 1):
            ws.cell(row=4 + i, column=col, value=val)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("\nUsage:")
        print("  # Shopify only (public store)")
        print("  python brand_audit_extractor.py https://www.skinq.com/products.json")
        print()
        print("  # Shopify Admin API with access token")
        print("  python brand_audit_extractor.py https://skinq.myshopify.com/admin \\")
        print("    --access-token shpat_xxxxxxxxxxxxxxxxxxxx")
        print()
        print("  # Shopify Admin API with Client ID + Secret")
        print("  python brand_audit_extractor.py https://skinq.myshopify.com/admin \\")
        print("    --client-id YOUR_CLIENT_ID \\")
        print("    --client-secret YOUR_CLIENT_SECRET")
        print()
        print("  # Full: Shopify + GA4 + GSC")
        print("  python brand_audit_extractor.py https://store.myshopify.com/admin \\")
        print("    --access-token shpat_xxxxxxxxxxxxxxxxxxxx \\")
        print("    --ga4-property 526018830 \\")
        print("    --ga4-credentials ./ga4_credentials.json \\")
        print("    --gsc-site \"https://www.example.com/\" \\")
        print("    --gsc-credentials ./ga4_credentials.json \\")
        print("    --output ./sheets/")
        print()
        sys.exit(1)

    # Parse arguments
    url = sys.argv[1]
    output_dir = "./"
    client_id = client_secret = access_token = None
    ga4_property = ga4_credentials = gsc_site = gsc_credentials = None
    ga4_only = False

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == "--output" and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]
        elif arg == "--client-id" and i + 1 < len(sys.argv):
            client_id = sys.argv[i + 1]
        elif arg == "--client-secret" and i + 1 < len(sys.argv):
            client_secret = sys.argv[i + 1]
        elif arg == "--access-token" and i + 1 < len(sys.argv):
            access_token = sys.argv[i + 1]
        elif arg == "--ga4-property" and i + 1 < len(sys.argv):
            ga4_property = sys.argv[i + 1]
        elif arg == "--ga4-credentials" and i + 1 < len(sys.argv):
            ga4_credentials = sys.argv[i + 1]
        elif arg == "--gsc-site" and i + 1 < len(sys.argv):
            gsc_site = sys.argv[i + 1]
        elif arg == "--gsc-credentials" and i + 1 < len(sys.argv):
            gsc_credentials = sys.argv[i + 1]
        elif arg == "--ga4-only":
            ga4_only = True

    os.makedirs(output_dir, exist_ok=True)

    if not url.startswith("http"):
        url = "https://" + url

    # Handle admin.shopify.com/store/HANDLE format
    if "admin.shopify.com/store/" in url:
        store = url.split("admin.shopify.com/store/")[1].split("/")[0].split("?")[0]
        base_url = f"https://{store}.myshopify.com"
    else:
        domain = url.split("//")[1].split("/")[0]
        store = domain.replace("www.", "")
        base_url = f"https://{store}"
    today = datetime.now().strftime("%d %b %Y").upper()
    is_admin = "/admin" in url or access_token or client_id

    # GA4-only mode: skip Shopify
    if ga4_only:
        products = []
        brand_name = store.split(".")[0].replace("-", " ").title()
        print(f"\n{'=' * 60}")
        print(f"Brand Audit Extractor (GA4-Only Mode)")
        print(f"{'=' * 60}")
        print(f"  Brand:   {brand_name}")
        print(f"  Mode:    GA4 Only (Shopify skipped)")
        print(f"  Store:   {store}")

    print(f"\n{'=' * 60}")
    print(f"Brand Audit Extractor (Universal)")
    print(f"{'=' * 60}")
    print(f"  Store:   {store}")
    print(f"  Mode:    {'Admin API + GraphQL' if is_admin else 'Public JSON'}")

    # ── Fetch Shopify products ──────────────────────────────────────────────
    products = []
    field_map = {}

    if not ga4_only:
        try:
            if client_id and client_secret:
                access_token = get_shopify_token(base_url, client_id, client_secret)

            if access_token:
                # Phase 1: Discover schema
                schema = discover_metafield_schema(base_url, access_token)
                field_map = schema.get("field_map", {})

                # Phase 2: Extract all products
                raw_products = fetch_products_graphql(base_url, access_token)

                # Phase 3: Normalize each product
                products = [extract_product_universal(p, base_url, "graphql", field_map) for p in raw_products]

                # Get brand name from first product's vendor
                if products:
                    brand_name = products[0].get("_vendor", "")
            else:
                if "/products.json" not in url:
                    url = url.rstrip("/") + "/products.json"
                raw_products = fetch_products_public_json(url)
                products = [extract_product_universal(p, base_url, "public", field_map) for p in raw_products]
                if products:
                    brand_name = products[0].get("_vendor", "")

        except requests.exceptions.HTTPError as e:
            print(f"  HTTP Error: {e}")
            sys.exit(1)
        except requests.exceptions.ConnectionError:
            print(f"  Connection Error. Check the URL.")
            sys.exit(1)
        except Exception as e:
            print(f"  Shopify Error: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

        if not brand_name:
            brand_name = store.split(".")[0].replace("-", " ").title()

        print(f"  Brand:   {brand_name}")
    else:
        brand_name = store.split(".")[0].replace("-", " ").title()

    # ── Fetch GA4 data ──────────────────────────────────────────────────────
    ga4_data = None
    if ga4_property and ga4_credentials:
        if not os.path.exists(ga4_credentials):
            print(f"\n  WARNING: GA4 credentials not found at {ga4_credentials}")
        else:
            print(f"\n  Fetching GA4 data from property {ga4_property}...")
            try:
                ga4_data = fetch_ga4(ga4_property, ga4_credentials)
                s = ga4_data["summary"]
                print(f"  GA4: {s['total_sessions']:,} sessions, {s['total_users']:,} users (30 days)")
            except Exception as e:
                print(f"  GA4 Error: {e}")

    # ── Fetch GSC data ─────────────────────────────────────────────────────
    gsc_data = None
    gsc_cred_file = gsc_credentials or (ga4_credentials if gsc_site else None)
    if gsc_site and gsc_cred_file:
        if not os.path.exists(gsc_cred_file):
            print(f"\n  WARNING: GSC credentials not found at {gsc_cred_file}")
        else:
            print(f"\n  Fetching GSC data from site {gsc_site}...")
            try:
                gsc_data = fetch_gsc(gsc_site, gsc_cred_file)
                s = gsc_data["summary"]
                print(f"  GSC: {s['total_clicks']:,} clicks, {s['total_impressions']:,} impressions (90 days)")
            except Exception as e:
                print(f"  GSC Error: {e}")

    if not products and not ga4_data and not gsc_data:
        print("  No data found. Exiting.")
        sys.exit(1)

    # ── GA4-only skip: ensure at least one source succeeded ────────────────
    if ga4_only and not ga4_data and not gsc_data:
        print("  GA4-only mode: no GA4 or GSC data found. Exiting.")
        sys.exit(1)
        print("  No data found. Exiting.")
        sys.exit(1)

    # ── GA4-only skip: ensure at least one source succeeded ────────────────
    if ga4_only and not ga4_data and not gsc_data:
        print("  GA4-only mode: no GA4 or GSC data found. Exiting.")
        sys.exit(1)

    # ── Build Excel ──────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{store.replace('.', '_')}_audit_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    ga4_count = 6 if ga4_data else 0
    gsc_count = 4 if gsc_data else 0
    print(f"\n  Building Excel (1 summary + {len(products)} product sheets + {ga4_count} GA4 + {gsc_count} GSC)...")
    build_excel(products, brand_name, base_url, today, output_path,
                ga4_data=ga4_data, ga4_property=ga4_property or "",
                gsc_data=gsc_data, gsc_site=gsc_site or "")

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print(f"Done!")
    print(f"  Brand:      {brand_name}")
    if products:
        active = sum(1 for p in products if p.get("_status") == "ACTIVE")
        draft = sum(1 for p in products if p.get("_status") == "DRAFT")
        print(f"  Products:   {len(products)} total ({active} active, {draft} draft)")
        rated = sum(1 for p in products if p.get("Rating", "—") != "—")
        print(f"  With ratings: {rated}")
    if ga4_data:
        s = ga4_data["summary"]
        print(f"  GA4:        {s['total_sessions']:,} sessions ({s['date_range']})")
    if gsc_data:
        s = gsc_data["summary"]
        print(f"  GSC:        {s['total_clicks']:,} clicks ({s['date_range']})")
    print(f"  Output:     {output_path}")
    print()


if __name__ == "__main__":
    main()